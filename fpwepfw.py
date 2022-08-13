import openpyxl

file_name = ['', '' , 'ru', 'en', 'ar', 'cs', 'de', 'es', 'fi', 'fr', 'it', 'ja', 'pl', 'pt_br', 'th', 'tr', 'vi', 'zh-cn', 'zh_tw', 'ko']
workbook = openpyxl.load_workbook('')

product_name = ''

sheets = workbook.sheetnames
sheet = workbook.active

read_data = ''

for column in range(2, 18):
    with open(r'blueprint.po', "r+", encoding="utf-8") as f:
        read_data = f.read()
    # for column in range(2, 4):
    #     read_data += '<p>{0}</p>\n'.format(str(sheet.cell(column=row, row=column).value))
    my_file = open(f"{file_name[column]}.po", "w+", encoding='utf-8')
    read_data = read_data.replace('%product_name%', str(sheet.cell(column=column, row=2).value))
    read_data = read_data.replace('%product_description%', str(sheet.cell(column=column, row=3).value))
    read_data = read_data.replace('%package_content%', str(sheet.cell(column=column, row=4).value))
    my_file.write(read_data)
    my_file.close()

print(read_data)
